# *******************************
# NOTES
#
# - This code was developed & tested using Python version 3.7.4
# - On my machine, this code takes roughly 16 minutes to run
# - The code can be run regardless of whether or not the keyword.txt / distance.xlsx
#   files already exist in the current directory, but it will overwrite
#   them if they do anyway.
# - However, a valid keyword.xlsx file must exist in the current directory.
# - The code will generate to the current directory, a keyword.txt file for 
#   every keyword, a distance.xlsx file, and two .png files that represent
#   visual graphs generated in seaborn.
# *******************************


import urllib.request
import urllib.parse
from bs4 import BeautifulSoup
from bs4.element import Comment
from openpyxl import load_workbook
import seaborn as sns
from re import sub
from gensim.utils import simple_preprocess
import gensim.downloader
from gensim.corpora import Dictionary
from gensim.models import TfidfModel, WordEmbeddingSimilarityIndex
from gensim.similarities import SparseTermSimilarityMatrix, SoftCosineSimilarity
import numpy as np
import nltk
nltk.download('stopwords')
from nltk.corpus import stopwords
from shutil import copyfile
import pandas as pd


# *******************************
# Initial Setup
# *******************************

print('Running... this code may take a minimum of 16 minutes to execute.')

# Read in keywords from excel file
keywords = []
keywords_check = []
wb = load_workbook(filename = 'keywords.xlsx')
worksheet = wb.worksheets[0]
for row in range(2, worksheet.max_row+1):
    cell_name = "{}{}".format('A', row)
    keywords.append(worksheet[cell_name].value)
for column in "BCDEFGHIJK":
    cell_name = "{}{}".format(column, 1)
    keywords_check.append(worksheet[cell_name].value)

# Check that the excel file is a valid matrix
if keywords != keywords_check:
    print("""Error: The keywords.xlsx file does not contain a valid matrix. Ensure that the values of
    the cells in row 1, from column B onwards are the exact same and in the exact same order as the
    values of the cells in column A from row 2 downwards.""")
    exit()

# *******************************
# Goals 1 & 2
# *******************************

def remove_misc_articles(articles):
    misc = ['/localnews', '/help']
    final_articles = []
    for a in articles:
        to_use = True
        for m in misc:
            if m in a:
                to_use = False
                break
        if to_use:
            final_articles.append(a)
    return final_articles

def check_article_content(element):
    par = element.parent
    if par.name == 'a' and element == 'About sharing':
        return False
    if par.name == 'h2' and element in ['More on this story', 'Related Topics', 'Related Internet Links', 'Around the BBC']:
        return False
    if par.name in ['span','div','dt','button','li','style']:
        return False
    if par.parent.name in ['li','cite','footer']:
        return False
    if isinstance(element, Comment):
        return False
    return True

url = 'https://bbc.co.uk/search'
num_articles_for_keyword = []
for word in keywords:
    print('Now searching for articles by keyword: ' + word)
    overall_articles = []
    page_number = 0
    while True:
        page_number += 1
        articles = []
        values = {'q' : word, 'page' : page_number}
        data = urllib.parse.urlencode(values)
        req = urllib.request.Request(url+'?'+data)
        resp = urllib.request.urlopen(req)
        respData = resp.read()
        soup = BeautifulSoup(respData, 'html.parser')
        for link in soup.find_all('a'):
            if '/news/' in link.get('href'): # ensures we are considering a News article only
                articles.append(link.get('href'))
        articles = remove_misc_articles(articles)
        if len(articles) == 0:
            break
        for article in articles:
            if len(overall_articles) < 100:
                overall_articles.append(article)
            else:
                break
        if len(overall_articles) == 100:
            break
    num_articles_for_keyword.append(len(overall_articles))
    print(str(len(overall_articles)) + ' articles found for keyword: ' + word)
    print('Now writing article contents to file...')
    f = open(word + '.txt', 'w')
    for article in overall_articles:
        req = urllib.request.Request(article)
        resp = urllib.request.urlopen(req)
        respData = resp.read()
        soup = BeautifulSoup(respData, 'html.parser')
        try:
            texts = soup.article.find_all(string=True)
        except:
            print('Could not grab article contents for: ' + article + ' - moving onto next article')
            continue
        texts_of_note = filter(check_article_content, texts)
        f.write(u" ".join(t.strip() for t in texts_of_note))
    print('File successfully complete for keyword: ' + word)

# *******************************
# Goal 3
# *******************************

print('Now calculating semantic distances...')

def clean_and_strip(document):
    document = sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', ' url_token ', document)
    document = simple_preprocess(document, min_len = 0, max_len = float('inf'))
    word_list = []
    for w in document:
        if w not in stopwords.words('english'):
            word_list.append(w)
    return word_list

try:
    wb = load_workbook(filename = 'distance.xlsx')
except:
    copyfile('keywords.xlsx', 'distance.xlsx')
    wb = load_workbook(filename = 'distance.xlsx')
worksheet = wb.worksheets[0]
# Enter a value 1 across the matrix diagonal, because 2 identical keywords have a semantic distance of 1
for i in range(2, (len(keywords)+2)):
    worksheet.cell(row=i, column=i).value = 1

processed_docs = []
for keyword in keywords:
    orig = open((keyword + '.txt'), "r").read()
    processed_docs.append(clean_and_strip(orig))

similarity_idx = WordEmbeddingSimilarityIndex(gensim.downloader.load("glove-wiki-gigaword-50"))
word_dictionary = Dictionary(processed_docs)
tf_idf = TfidfModel(dictionary = word_dictionary)
term_similarity_matrix = SparseTermSimilarityMatrix(similarity_idx, word_dictionary, tf_idf)

done_keywords = []
for k in keywords:
    corpus = []
    corpus_keywords = []
    query = processed_docs[keywords.index(k)]
    done_keywords.append(k)
    for k2 in keywords:
        if k2 == k:
            continue
        else:
            corpus_keywords.append(k2)
            corpus.append(processed_docs[keywords.index(k2)])

    query_tf = tf_idf[word_dictionary.doc2bow(query)]
    index = SoftCosineSimilarity(tf_idf[[word_dictionary.doc2bow(doc) for doc in corpus]], term_similarity_matrix)
    doc_similarity_scores = index[query_tf]

    for i in np.argsort(doc_similarity_scores)[::-1]:
        if corpus_keywords[i] in done_keywords:
            continue
        worksheet.cell(row=(keywords.index(corpus_keywords[i])+2), column=(keywords.index(k)+2)).value = doc_similarity_scores[i]
        worksheet.cell(row=(keywords.index(k)+2), column=(keywords.index(corpus_keywords[i])+2)).value = doc_similarity_scores[i]

wb.save('distance.xlsx')
print('Semantic Distances have been calculated and stored to distance.xlsx')


# *******************************
# Goal 4
# *******************************

# Number of articles per keyword bar chart
ax1 = sns.barplot(x=keywords, y=num_articles_for_keyword)
ax1.set_xticklabels(ax1.get_xticklabels(), fontsize=5)
fig1 = ax1.get_figure()
fig1.savefig('article_count.png')
ax1.get_figure().clf()

# Semantic Distance Heatmap
df = pd.read_excel('distance.xlsx')
df.set_index(['Keywords'], inplace=True)
ax2 = sns.heatmap(df)
fig2 = ax2.get_figure()
fig2.savefig('heatmap.png', bbox_inches = 'tight')

print('Seaborn graphs have been generated')
print('Execution has finished')
